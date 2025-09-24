#!/usr/bin/env python3
"""
Cross-platform Excel -> PDF converter that preserves:
- Gridlines
- Row/Column headings (A,B,C... / 1,2,3...)
- Reasonable fit-to-page defaults
- Auto-fit to prevent #### truncation

Usage:
  python excel2pdf.py <file_or_folder> [--out <output_folder>] [--landscape] [--fitwide N]

Notes:
- If Excel is available (macOS/Windows), uses xlwings for best fidelity (autofit via Excel).
- Otherwise, falls back to LibreOffice headless; before that, we widen columns via openpyxl.
"""

import sys, os, subprocess, platform, tempfile, argparse
from pathlib import Path

# -------------------------
# openpyxl autofit helpers
# -------------------------
def _approx_display_len(cell):
    """Estimate display width of a cell's value based on its number format."""
    v = cell.value
    if v is None:
        return 0
    try:
        # crude number formatting approximation
        if isinstance(v, (int, float)):
            nf = (cell.number_format or "").lower()
            # decide decimals
            if any(tok in nf for tok in ["0.00", "#,##0.00", "0.0"]):
                s = f"{v:,.2f}"
            elif any(tok in nf for tok in ["0.000", "0.0000"]):
                s = f"{v:,.3f}"
            else:
                s = f"{v:,}"
            return len(s)
        return len(str(v))
    except Exception:
        return len(str(v))

def autofit_columns_openpyxl(ws, min_width=8, padding=2, max_width=120):
    from openpyxl.utils import get_column_letter
    lengths = {}
    # consider used range only, iterate all cells with values
    for row in ws.iter_rows(values_only=False):
        for cell in row:
            if cell.value is None:
                continue
            col_idx = cell.column
            L = _approx_display_len(cell)
            if L > 0:
                if col_idx not in lengths or L > lengths[col_idx]:
                    lengths[col_idx] = L
    # convert to Excel width units (roughly the number of '0' chars in Calibri 11)
    for col_idx, L in lengths.items():
        letter = get_column_letter(col_idx)
        target = max(min_width, min(max_width, L + padding))
        ws.column_dimensions[letter].width = target

def ensure_print_settings(xlsx_path: Path, landscape: bool, fitwide: int) -> Path:
    from openpyxl import load_workbook
    wb = load_workbook(xlsx_path)
    for ws in wb.worksheets:
        # Widen columns so LibreOffice won't produce ####
        try:
            autofit_columns_openpyxl(ws)
        except Exception:
            pass

        # Show gridlines & headings
        ws.print_options.gridLines = True
        ws.print_options.headings = True

        # Fit: N pages wide, as many tall as needed
        ws.page_setup.fitToWidth = max(1, int(fitwide))
        ws.page_setup.fitToHeight = 0

        # Optional: orientation
        if landscape:
            ws.page_setup.orientation = 'landscape'
        else:
            ws.page_setup.orientation = 'portrait'

    # Save to a temp copy so we never touch the original
    tmp = Path(tempfile.gettempdir()) / (xlsx_path.stem + "_print_ready.xlsx")
    wb.save(tmp)
    return tmp

# -------------------------
# Excel (xlwings) exporter
# -------------------------
def try_export_with_excel_via_xlwings(xlsx_path: Path, pdf_path: Path, landscape: bool, fitwide: int) -> bool:
    try:
        import xlwings as xw
    except Exception:
        return False
    try:
        app = xw.App(visible=False)
        try:
            wb = xw.Book(str(xlsx_path))

            # Auto-fit and page setup on each sheet
            for sht in wb.sheets:
                # Autofit columns and rows using Excel's native logic
                try:
                    sht.autofit('c')  # columns
                    sht.autofit('r')  # rows
                except Exception:
                    pass

                ps = sht.api.PageSetup
                ps.PrintGridlines = True
                ps.PrintHeadings = True
                ps.Zoom = False
                ps.FitToPagesWide = max(1, int(fitwide))
                ps.FitToPagesTall = False
                ps.Orientation = 2 if landscape else 1  # 1=Portrait, 2=Landscape

            # Export the entire workbook as a single PDF
            wb.to_pdf(str(pdf_path))  # xlwings 0.30+
            wb.close()
            return True
        finally:
            app.kill()
    except Exception:
        return False

# -------------------------
# LibreOffice exporter
# -------------------------
def try_export_with_libreoffice(xlsx_path: Path, pdf_path: Path) -> bool:
    # Use soffice headless to convert; relies on print settings we wrote into the file.
    soffice = "soffice"
    if platform.system() == "Darwin":  # macOS default install path
        mac_path = "/Applications/LibreOffice.app/Contents/MacOS/soffice"
        if os.path.exists(mac_path):
            soffice = mac_path
    cmd = [
        soffice, "--headless", "--norestore", "--nolockcheck",
        "--convert-to", "pdf:calc_pdf_Export",
        "--outdir", str(pdf_path.parent),
        str(xlsx_path)
    ]
    try:
        subprocess.run(cmd, check=True, stdout=subprocess.PIPE, stderr=subprocess.PIPE)
        # LibreOffice outputs to outdir with same stem; rename if needed
        produced = xlsx_path.with_suffix(".pdf").name
        src = pdf_path.parent / produced
        if src != pdf_path and src.exists():
            src.replace(pdf_path)
        return pdf_path.exists()
    except Exception:
        return False

# -------------------------
# Driver
# -------------------------
def convert_one(xlsx: Path, out_pdf: Path, landscape: bool, fitwide: int):
    # 1) Bake the print settings & our openpyxl autofit into a temp copy
    prepared = ensure_print_settings(xlsx, landscape=landscape, fitwide=fitwide)

    # 2) Prefer Excel (macOS/Windows), else LibreOffice
    used_excel = try_export_with_excel_via_xlwings(prepared, out_pdf, landscape=landscape, fitwide=fitwide)
    if not used_excel:
        used_libre = try_export_with_libreoffice(prepared, out_pdf)
        if not used_libre:
            raise RuntimeError("Could not export to PDF via Excel or LibreOffice.")

def is_excel_file(p: Path) -> bool:
    return p.suffix.lower() in {".xlsx", ".xlsm", ".xls"}

def parse_args():
    ap = argparse.ArgumentParser(description="Excel → PDF (gridlines + headings) with autofit.")
    ap.add_argument("target", help="Input Excel file or folder")
    ap.add_argument("--out", dest="out_dir", help="Output folder for PDFs")
    ap.add_argument("--landscape", action="store_true", help="Force landscape orientation")
    ap.add_argument("--fitwide", type=int, default=1, help="Fit to N pages wide (default: 1)")
    return ap.parse_args()

def main():
    args = parse_args()

    target = Path(args.target).expanduser().resolve()
    out_dir = Path(args.out_dir).expanduser().resolve() if args.out_dir else (target if target.is_dir() else target.parent)
    out_dir.mkdir(parents=True, exist_ok=True)

    if target.is_dir():
        files = [p for p in sorted(target.iterdir()) if p.is_file() and is_excel_file(p)]
    else:
        if not is_excel_file(target):
            print("Error: input is not an Excel file (.xlsx/.xlsm/.xls).")
            sys.exit(2)
        files = [target]

    failures = []
    for f in files:
        pdf_path = out_dir / (f.stem + ".pdf")
        try:
            convert_one(f, pdf_path, landscape=args.landscape, fitwide=args.fitwide)
            print(f"✓ {f.name} -> {pdf_path}")
        except Exception as e:
            print(f"✗ {f.name} failed: {e}")
            failures.append((f, e))

    if failures:
        sys.exit(3)

if __name__ == "__main__":
    main()
